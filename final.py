import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import io, math, datetime as dt, re
import openai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ====== PDF imports (reportlab) ======
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, LongTable, TableStyle, Paragraph, Spacer, PageBreak, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

# =====================================================
# 1. PAGE SETUP & DATA LOADING
# =====================================================
st.set_page_config(page_title="Jarir Quant Analyzer", layout="wide")
st.title("Jarir Quant Analysis")

uploaded_file = st.file_uploader("Please upload the raw quant file to begin.", type=["csv", "xlsx"])

if uploaded_file is None:
    st.info("Please upload the raw quant file to begin.")
    st.stop()

@st.cache_data
def load_and_clean_base(file):
    df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
    df.columns = df.columns.str.strip()
    if "Data" in df.columns: df.rename(columns={"Data": "Date"}, inplace=True)
    df["Date"] = pd.to_datetime(df["Date"], format='%b-%y', errors="coerce")
    if "Year" in df.columns:
        df['Date'] = df.apply(lambda x: x['Date'].replace(year=int(x['Year'])) if pd.notnull(x['Date']) else x['Date'], axis=1)
    mgr_cols = [c for c in df.columns if c not in ["Date", "Year"]]
    last_idx = df[mgr_cols].dropna(how='all').index.max()
    return df.iloc[:last_idx + 1].sort_values("Date").reset_index(drop=True)

df_raw = load_and_clean_base(uploaded_file)

# =====================================================
# 2. SIDEBAR SETTINGS
# =====================================================
st.sidebar.header("Analysis Settings")
available_dates = df_raw["Date"].dropna().sort_values().unique()
date_labels = [d.strftime('%b-%Y') for d in available_dates]

col_s1, col_s2 = st.sidebar.columns(2)
with col_s1:
    start_label = st.selectbox("Start Month", options=date_labels, index=0)
with col_s2:
    end_label = st.selectbox("End Month", options=date_labels, index=len(date_labels)-1)

start_dt = pd.to_datetime(start_label, format='%b-%Y')
end_dt   = pd.to_datetime(end_label,   format='%b-%Y')
df_filtered = df_raw[(df_raw["Date"] >= start_dt) & (df_raw["Date"] <= end_dt)].reset_index(drop=True)

rfr_target = "13 Wk US Treasury Bills"
default_rfr_idx = df_filtered.columns.get_loc(rfr_target) if rfr_target in df_filtered.columns else 0
rfr_col = st.sidebar.selectbox("Risk-Free Rate Column", options=df_filtered.columns, index=default_rfr_idx)

all_mgrs = [c for c in df_filtered.columns if c not in ["Date", "Year", rfr_col]]
manager_cols = st.sidebar.multiselect("Select Managers", options=all_mgrs, default=all_mgrs[:5])
if not manager_cols:
    st.warning("Please select managers.")
    st.stop()


# =====================================================
# 3. DATA CLEANING
# =====================================================
cleaned_df = df_filtered.copy()
for col in manager_cols + [rfr_col]:
    s = cleaned_df[col].astype(str).str.strip()
    is_pct = s.str.contains("%", regex=False)
    s = s.replace(["", "nan", "-", "–", " - "], np.nan)\
         .str.replace("%", "", regex=False)\
         .str.replace(r"\((.*?)\)", r"-\1", regex=True)\
         .astype(float)
    s.loc[is_pct] = s.loc[is_pct] / 100
    cleaned_df[col] = s

# =====================================================
# 4. CALCULATION HELPERS
# =====================================================
def get_cap(s_m, s_b):
    combined = pd.DataFrame({'m': s_m, 'b': s_b}).dropna()
    up = combined[combined['b'] > 0]; dn = combined[combined['b'] < 0]
    u_cap = (np.prod(1+up['m'])/np.prod(1+up['b']))*100 if not up.empty else np.nan
    d_cap = (np.prod(1+dn['m'])/np.prod(1+dn['b']))*100 if not dn.empty else np.nan
    return u_cap, d_cap

def slice_series(series, years=None):
    s = series.dropna()
    if years is None: return s
    months = int(years * 12)
    return s.tail(months) if len(s) >= months else None

def get_ann_ret(s):
    if s is None or len(s) < 12: return np.nan
    return (np.prod(1 + s)) ** (1 / (len(s) / 12)) - 1

def get_ann_vol(s):
    if s is None or len(s) < 2: return np.nan
    return s.std(ddof=1) * np.sqrt(12)

def get_downside_dev(s):
    if s is None or len(s) < 12: return np.nan
    return np.sqrt(np.mean(np.minimum(s, 0) ** 2)) * np.sqrt(12)

def get_upside_dev(s):
    if s is None or len(s) < 12: return np.nan
    return np.sqrt(np.mean(np.maximum(s, 0) ** 2)) * np.sqrt(12)

def get_max_drawdown(s):
    if s is None or len(s) == 0: return np.nan
    cum = (1 + s).cumprod()
    return ((cum - cum.cummax()) / cum.cummax()).min()

# =====================================================
# 5. ANALYTICS ENGINE  (extended horizons: 1–10, 15, 20)
# =====================================================
horizons   = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 20]
h_labels   = [f"{y} Year" for y in horizons] + ["Since Inception"]
metrics    = ["Annualized Return (%)", "Annualized Volatility (%)",
              "Upward Deviation (%)", "Downward Deviation (%)",
              "Sharpe Ratio", "Sortino Ratio", "Max Drawdown (%)"]
results    = {m: pd.DataFrame(index=h_labels, columns=manager_cols) for m in metrics}

for yrs, lbl in zip(horizons + [None], h_labels):
    for mgr in manager_cols:
        s_m  = slice_series(cleaned_df[mgr],   yrs)
        s_rf = slice_series(cleaned_df[rfr_col], yrs)
        if s_m is None or s_rf is None: continue

        ann_ret  = get_ann_ret(s_m)
        ann_rfr  = get_ann_ret(s_rf)
        ann_vol  = get_ann_vol(s_m)
        ann_ddev = get_downside_dev(s_m)
        ann_udev = get_upside_dev(s_m)
        exc_ret  = ann_ret - ann_rfr

        results["Annualized Return (%)"].at[lbl, mgr]  = ann_ret  * 100  if not np.isnan(ann_ret)  else np.nan
        results["Annualized Volatility (%)"].at[lbl, mgr] = ann_vol * 100 if not np.isnan(ann_vol) else np.nan
        results["Upward Deviation (%)"].at[lbl, mgr]   = ann_udev * 100  if not np.isnan(ann_udev) else np.nan
        results["Downward Deviation (%)"].at[lbl, mgr] = ann_ddev * 100  if not np.isnan(ann_ddev) else np.nan
        results["Sharpe Ratio"].at[lbl, mgr]  = exc_ret / ann_vol  if ann_vol  > 0 else np.nan
        results["Sortino Ratio"].at[lbl, mgr] = exc_ret / ann_ddev if ann_ddev > 0 else np.nan
        results["Max Drawdown (%)"].at[lbl, mgr] = get_max_drawdown(s_m) * 100

# =====================================================
# 6. TAB STRUCTURE
# =====================================================
tab1, tab2 = st.tabs([" Quant Analysis", " Strategic Visualization"])

with tab1:
    # =====================================================
    # 6. PERFORMANCE SNAPSHOT + KPI BANNER
    # =====================================================

    # ── KPI Banner Cards with selectable horizon ──
    st.markdown("### Performance Summary")
    banner_horizon = st.selectbox("Select Horizon for Summary Cards", options=h_labels, index=h_labels.index("Since Inception"), key="banner_h")

    si_ret    = results["Annualized Return (%)"].loc[banner_horizon].apply(pd.to_numeric, errors='coerce')
    si_vol    = results["Annualized Volatility (%)"].loc[banner_horizon].apply(pd.to_numeric, errors='coerce')
    si_sharpe = results["Sharpe Ratio"].loc[banner_horizon].apply(pd.to_numeric, errors='coerce')
    si_sort   = results["Sortino Ratio"].loc[banner_horizon].apply(pd.to_numeric, errors='coerce')
    si_dd     = results["Max Drawdown (%)"].loc[banner_horizon].apply(pd.to_numeric, errors='coerce')

    # Helper: get ann return for a manager at the banner horizon
    def _mgr_ret(mgr_name):
        try:
            v = float(results["Annualized Return (%)"].at[banner_horizon, mgr_name])
            return f"{v:+.2f}%" if not pd.isna(v) else ""
        except: return ""

    _card_bg  = ["#1C3A5E","#1A4731","#3D1515","#1A3A3A","#3D2010","#2D1F3D","#1A3545","#1E3A1E"]
    _card_acc = ["#4A9EDB","#52C788","#E07070","#4EC9B0","#E8915A","#9B72CF","#5BC4E8","#7DC87D"]

    _bm = si_sharpe.idxmax()
    _hr = si_ret.idxmax()
    _lr = si_ret.idxmin()
    _ld = si_dd.idxmax()
    _hd = si_dd.idxmin()
    _hv = si_vol.idxmax()
    _lv = si_vol.idxmin()
    _bs = si_sort.idxmax()

    banner_cards = [
        ("BEST SHARPE",         _bm, f"{si_sharpe.max():.2f}", _card_bg[0], _card_acc[0]),
        ("HIGHEST ANN. RETURN", _hr, f"{si_ret.max():.2f}%", _card_bg[1], _card_acc[1]),
        ("LOWEST ANN. RETURN",  _lr, f"{si_ret.min():.2f}%", _card_bg[2], _card_acc[2]),
        ("LOWEST DRAWDOWN",     _ld, f"{si_dd.max():.2f}%", _card_bg[3], _card_acc[3]),
        ("HIGHEST DRAWDOWN",    _hd, f"{si_dd.min():.2f}%", _card_bg[4], _card_acc[4]),
        ("HIGHEST VOLATILITY",  _hv, f"{si_vol.max():.2f}%", _card_bg[5], _card_acc[5]),
        ("LOWEST VOLATILITY",   _lv, f"{si_vol.min():.2f}%",  _card_bg[6], _card_acc[6]),
        ("BEST SORTINO",        _bs, f"{si_sort.max():.2f}",  _card_bg[7], _card_acc[7]),
    ]

    st.markdown("""
    <style>
    .banner-grid{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px}
    .banner-card{
        flex:1 1 155px;border-radius:12px;padding:16px 12px 0 12px;text-align:center;
        box-shadow:0 4px 14px rgba(0,0,0,.30);border:1px solid rgba(255,255,255,.08);
        display:flex;flex-direction:column;justify-content:space-between;min-height:130px;
    }
    .banner-card .label{font-size:9.5px;font-weight:700;opacity:.72;margin-bottom:8px;
        letter-spacing:.9px;text-transform:uppercase}
    .banner-card .mgr-row{display:flex;align-items:center;justify-content:center;gap:6px;margin-bottom:4px}
    .banner-card .manager{font-size:13px;font-weight:800;opacity:.97}
    .banner-card .ret-badge{font-size:10.5px;font-weight:700;padding:2px 8px;border-radius:20px;
        background:rgba(255,255,255,0.18);letter-spacing:.2px}
    .banner-card .value{font-size:24px;font-weight:900;letter-spacing:-.5px;margin:4px 0 10px}
    .banner-card .colour-bar{height:5px;border-radius:0 0 12px 12px;
        margin:0 -12px;width:calc(100% + 24px);flex-shrink:0}
    </style>""", unsafe_allow_html=True)

    cards_html = '<div class="banner-grid">'
    for lbl_b, mgr_b, val_b, bg_col, acc_col in banner_cards:
        cards_html += (
            f'<div class="banner-card" style="background:{bg_col};color:white;">'
            f'<div class="label">{lbl_b}</div>'
            f'<div class="mgr-row">'
            f'<span class="manager">{mgr_b}</span>'
            f'</div>'
            f'<div class="value" style="color:{acc_col};">{val_b}</div>'
            f'<div class="colour-bar" style="background:{acc_col};opacity:0.85;"></div>'
            f'</div>'
        )
    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)
    st.divider()

    # =====================================================
    # 7. STYLE HELPER
    # =====================================================
    def style_df(df_in, pct=True):
        fmt = "{:.2f}%" if pct else "{:.2f}"
        return df_in.apply(pd.to_numeric).style \
            .map(lambda x: 'color:red;text-align:center;font-weight:normal'
                 if (pd.notna(x) and x < 0)
                 else ('text-align:center;color:black;font-weight:normal' if pd.notna(x) else '')) \
            .format(fmt, na_rep="") \
            .set_properties(**{'text-align': 'center'}) \
            .set_table_styles([
                {'selector': 'th.row_heading', 'props': [('font-weight','bold'),('color','black'),('text-align','left'),('white-space','nowrap')]},
                {'selector': 'th.col_heading', 'props': [('font-weight','bold'),('color','black'),('text-align','center')]},
                {'selector': 'th.blank',       'props': [('text-align','center')]},
            ])

    # =====================================================
    # 8. METRIC TABLES
    # =====================================================
    for m in metrics:
        is_p = "Ratio" not in m
        with st.expander(f"View: {m}", expanded=True):
            st.dataframe(style_df(results[m], pct=is_p), width="stretch", height=(len(h_labels)+1)*35+15)

    # =====================================================
    # 9. RISK–RETURN + UPSIDE–DOWNSIDE PLOTS
    # =====================================================
    st.divider()
    viz_h_options = [lbl for lbl in h_labels if lbl != "6 Year"]
    viz_h = st.selectbox("Select Horizon for Plots", options=viz_h_options, index=len(viz_h_options)-2)

    col_v1, col_v2 = st.columns(2)

    def smart_scatter(ax, x_vals, y_vals, labels, dot_color, xlabel, ylabel, reflines=None):
        """Scatter with repulsion-based label placement — no overlaps."""
        import numpy as np

        ax.scatter(x_vals, y_vals, color=dot_color, s=90, zorder=5)

        if not labels:
            ax.set_xlabel(xlabel, fontsize=9)
            ax.set_ylabel(ylabel, fontsize=9)
            ax.grid(True, ls=':', alpha=0.6)
            return

        x_arr = np.array(x_vals, dtype=float)
        y_arr = np.array(y_vals, dtype=float)
        n = len(labels)

        x_range = max(x_arr.max() - x_arr.min(), 1e-6)
        y_range = max(y_arr.max() - y_arr.min(), 1e-6)

        # Start label positions offset from dots
        lx = x_arr + x_range * 0.04
        ly = y_arr + y_range * 0.04

        # Iterative repulsion between labels
        for _ in range(300):
            for i in range(n):
                fx, fy = 0.0, 0.0
                for j in range(n):
                    if i == j: continue
                    dx = (lx[i] - lx[j]) / x_range
                    dy = (ly[i] - ly[j]) / y_range
                    dist2 = dx*dx + dy*dy + 1e-9
                    if dist2 < 0.08:
                        fx += dx / dist2
                        fy += dy / dist2
                # Also repel from original dot positions
                for j in range(n):
                    dx = (lx[i] - x_arr[j]) / x_range
                    dy = (ly[i] - y_arr[j]) / y_range
                    dist2 = dx*dx + dy*dy + 1e-9
                    if dist2 < 0.06:
                        fx += 0.5 * dx / dist2
                        fy += 0.5 * dy / dist2
                step = 0.0008
                lx[i] += fx * step * x_range
                ly[i] += fy * step * y_range

        for i, lbl in enumerate(labels):
            ax.annotate(
                lbl,
                xy=(x_arr[i], y_arr[i]),
                xytext=(lx[i], ly[i]),
                fontsize=8.5, fontweight='bold', color='#1a1a2e',
                bbox=dict(boxstyle='round,pad=0.3', fc='white', ec='#cccccc', alpha=0.85, lw=0.6),
                arrowprops=dict(arrowstyle='-', color='#999999', lw=0.8),
                zorder=10
            )

        ax.set_xlabel(xlabel, fontsize=9)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.grid(True, ls=':', alpha=0.6)
        if reflines:
            for axis, val, lw in reflines:
                if axis == 'h': ax.axhline(val, color='black', lw=lw)
                if axis == 'v': ax.axvline(val, color='black', lw=lw)

    with col_v1:
        st.write(f"**Risk-Return Plot ({viz_h})**")
        rr_ui = pd.DataFrame({
            "Return (%)":     results["Annualized Return (%)"].loc[viz_h],
            "Volatility (%)": results["Annualized Volatility (%)"].loc[viz_h]
        }).apply(pd.to_numeric).dropna()

        x_rr      = rr_ui["Volatility (%)"].values.astype(float)
        y_rr      = rr_ui["Return (%)"].values.astype(float)
        labels_rr = rr_ui.index.tolist()
        _n_rr     = len(labels_rr)

        # ── RdYlGn gradient by return (same as PDF / image) ──
        _rr_norm  = plt.Normalize(y_rr.min(), y_rr.max())
        _rr_cmap  = plt.cm.RdYlGn
        _rr_cols  = [_rr_cmap(_rr_norm(v)) for v in y_rr]

        # ── Figure: scatter area + colourbar + right-panel ──
        fig_rr = plt.figure(figsize=(11, 5.5))
        fig_rr.patch.set_facecolor('#F9FAFB')
        # axes: [left, bottom, width, height]
        ax_rr   = fig_rr.add_axes([0.07, 0.11, 0.60, 0.78])   # main scatter
        ax_cb   = fig_rr.add_axes([0.69, 0.11, 0.025, 0.78])  # colourbar
        ax_leg  = fig_rr.add_axes([0.74, 0.05, 0.25, 0.88])   # right panel

        ax_rr.set_facecolor('#FAFBFC')

        # ── Physics-repulsion label placement ──
        _lx = x_rr + (x_rr.max()-x_rr.min())*0.03
        _ly = y_rr + (y_rr.max()-y_rr.min())*0.04
        _xr = max(x_rr.max()-x_rr.min(), 1e-6)
        _yr = max(y_rr.max()-y_rr.min(), 1e-6)
        for _ in range(400):
            for _i in range(_n_rr):
                _fx,_fy = 0.,0.
                for _j in range(_n_rr):
                    if _i==_j: continue
                    _dx=(_lx[_i]-_lx[_j])/_xr; _dy=(_ly[_i]-_ly[_j])/_yr
                    _d2=_dx*_dx+_dy*_dy+1e-9
                    if _d2<0.09: _fx+=_dx/_d2; _fy+=_dy/_d2
                for _j in range(_n_rr):
                    _dx=(_lx[_i]-x_rr[_j])/_xr; _dy=(_ly[_i]-y_rr[_j])/_yr
                    _d2=_dx*_dx+_dy*_dy+1e-9
                    if _d2<0.07: _fx+=0.6*_dx/_d2; _fy+=0.6*_dy/_d2
                _lx[_i]+=_fx*0.0007*_xr; _ly[_i]+=_fy*0.0007*_yr

        # ── Plot dots + repulsion labels ──
        sc_rr = ax_rr.scatter(x_rr, y_rr, c=y_rr, cmap=_rr_cmap,
                               norm=_rr_norm, s=120, zorder=5,
                               edgecolors='white', linewidths=0.8)
        for k in range(_n_rr):
            ax_rr.annotate(
                labels_rr[k],
                xy=(x_rr[k], y_rr[k]), xytext=(_lx[k], _ly[k]),
                fontsize=8, fontweight='bold', color='#1a1a2e',
                bbox=dict(boxstyle='round,pad=0.28', fc='white',
                          ec='#CCCCCC', alpha=0.92, lw=0.7),
                arrowprops=dict(arrowstyle='-', color='#999999', lw=0.8),
                zorder=10
            )

        ax_rr.set_xlabel("Annualized Volatility (%)", fontsize=9)
        ax_rr.set_ylabel("Annualized Return (%)", fontsize=9)
        ax_rr.set_title(f"Risk–Return Analysis — {viz_h}", fontsize=12,
                        fontweight='bold', color='#1C2B4A', pad=10)
        ax_rr.grid(True, ls=':', alpha=0.4, color='#CCCCCC')
        ax_rr.spines[['top','right']].set_visible(False)

        # ── Colourbar ──
        import matplotlib.colorbar as mcb
        cb_rr = mcb.ColorbarBase(ax_cb, cmap=_rr_cmap, norm=_rr_norm,
                                  orientation='vertical')
        cb_rr.set_label("Return (%)", fontsize=8, color='#555')
        cb_rr.ax.tick_params(labelsize=7)

        # ── Right panel: managers sorted descending by return ──
        ax_leg.set_facecolor('#F9FAFB')
        ax_leg.axis('off')
        _sorted_idx = np.argsort(y_rr)[::-1]   # descending
        _panel_y    = 0.97
        _line_h     = 0.88 / _n_rr
        ax_leg.text(0.5, 1.01, "Return (%)", ha='center', va='bottom',
                    fontsize=8, fontweight='bold', color='#1C2B4A',
                    transform=ax_leg.transAxes)
        for _rank, _ki in enumerate(_sorted_idx):
            _col_dot = _rr_cols[_ki]
            _label   = labels_rr[_ki]
            _ret_val = y_rr[_ki]
            _row_y   = _panel_y - _rank * _line_h
            # coloured dot
            ax_leg.plot(0.06, _row_y, 'o', color=_col_dot, markersize=8,
                        transform=ax_leg.transAxes, clip_on=False)
            # manager name
            ax_leg.text(0.17, _row_y, _label, ha='left', va='center',
                        fontsize=7.5, fontweight='bold', color='#1a1a2e',
                        transform=ax_leg.transAxes)
            # return value right-aligned
            ax_leg.text(0.98, _row_y, f"{_ret_val:+.2f}%", ha='right', va='center',
                        fontsize=7.5, color=_col_dot, fontweight='bold',
                        transform=ax_leg.transAxes)

        st.pyplot(fig_rr)
        plt.close(fig_rr)



    with col_v2:
        b_cap = st.selectbox("Benchmark for Capture Scatter", options=manager_cols, index=len(manager_cols)-1)
        st.write(f"**Capture Matrix ({viz_h}) — Benchmark: {b_cap}**")
        caps_ui = []
        for m in manager_cols:
            y_c  = int(viz_h.split()[0]) if "Year" in viz_h else None
            s_m_ = cleaned_df[m].tail(y_c*12) if y_c else cleaned_df[m]
            s_b_ = cleaned_df[b_cap].tail(y_c*12) if y_c else cleaned_df[b_cap]
            u, d = get_cap(s_m_.dropna(), s_b_.dropna())
            caps_ui.append({"Manager": m, "Upside": u, "Downside": d})
        cap_ui_df = pd.DataFrame(caps_ui).set_index("Manager").dropna()

        x_cap      = cap_ui_df["Downside"].values.astype(float)
        y_cap      = cap_ui_df["Upside"].values.astype(float)
        labels_cap = cap_ui_df.index.tolist()
        _n_cap     = len(labels_cap)

        # ── RdYlGn gradient by upside capture ──
        _cap_norm = plt.Normalize(y_cap.min(), y_cap.max())
        _cap_cmap = plt.cm.RdYlGn
        _cap_cols = [_cap_cmap(_cap_norm(v)) for v in y_cap]

        # ── Figure: scatter + colourbar + right panel ──
        fig_cap = plt.figure(figsize=(11, 5.5))
        fig_cap.patch.set_facecolor('#F9FAFB')
        ax_cap  = fig_cap.add_axes([0.07, 0.11, 0.60, 0.78])
        ax_cbc  = fig_cap.add_axes([0.69, 0.11, 0.025, 0.78])
        ax_lc   = fig_cap.add_axes([0.74, 0.05, 0.25, 0.88])
        ax_cap.set_facecolor('#FAFBFC')

        # ── Physics-repulsion ──
        _lxc = x_cap + (x_cap.max()-x_cap.min())*0.03
        _lyc = y_cap + (y_cap.max()-y_cap.min())*0.04
        _xrc = max(x_cap.max()-x_cap.min(), 1e-6)
        _yrc = max(y_cap.max()-y_cap.min(), 1e-6)
        for _ in range(400):
            for _i in range(_n_cap):
                _fx,_fy = 0.,0.
                for _j in range(_n_cap):
                    if _i==_j: continue
                    _dx=(_lxc[_i]-_lxc[_j])/_xrc; _dy=(_lyc[_i]-_lyc[_j])/_yrc
                    _d2=_dx*_dx+_dy*_dy+1e-9
                    if _d2<0.09: _fx+=_dx/_d2; _fy+=_dy/_d2
                for _j in range(_n_cap):
                    _dx=(_lxc[_i]-x_cap[_j])/_xrc; _dy=(_lyc[_i]-y_cap[_j])/_yrc
                    _d2=_dx*_dx+_dy*_dy+1e-9
                    if _d2<0.07: _fx+=0.6*_dx/_d2; _fy+=0.6*_dy/_d2
                _lxc[_i]+=_fx*0.0007*_xrc; _lyc[_i]+=_fy*0.0007*_yrc

        ax_cap.scatter(x_cap, y_cap, c=y_cap, cmap=_cap_cmap,
                       norm=_cap_norm, s=120, zorder=5,
                       edgecolors='white', linewidths=0.8)
        for k in range(_n_cap):
            ax_cap.annotate(
                labels_cap[k],
                xy=(x_cap[k], y_cap[k]), xytext=(_lxc[k], _lyc[k]),
                fontsize=8, fontweight='bold', color='#1a1a2e',
                bbox=dict(boxstyle='round,pad=0.28', fc='white',
                          ec='#CCCCCC', alpha=0.92, lw=0.7),
                arrowprops=dict(arrowstyle='-', color='#999999', lw=0.8),
                zorder=10
            )

        ax_cap.axhline(100, color='#444444', lw=0.9, ls='--', zorder=4)
        ax_cap.axvline(100, color='#444444', lw=0.9, ls='--', zorder=4)
        ax_cap.set_xlabel("Downside Capture (%)", fontsize=9)
        ax_cap.set_ylabel("Upside Capture (%)", fontsize=9)
        ax_cap.set_title(f"Capture Matrix — {viz_h}  ·  Benchmark: {b_cap}",
                         fontsize=11, fontweight='bold', color='#1C2B4A', pad=10)
        ax_cap.grid(True, ls=':', alpha=0.4, color='#CCCCCC')
        ax_cap.spines[['top','right']].set_visible(False)

        # ── Colourbar ──
        import matplotlib.colorbar as mcb
        cb_cap = mcb.ColorbarBase(ax_cbc, cmap=_cap_cmap, norm=_cap_norm,
                                   orientation='vertical')
        cb_cap.set_label("Upside (%)", fontsize=8, color='#555')
        cb_cap.ax.tick_params(labelsize=7)

        # ── Right panel: managers sorted descending by upside capture ──
        ax_lc.set_facecolor('#F9FAFB')
        ax_lc.axis('off')
        _sorted_cap = np.argsort(y_cap)[::-1]
        _panel_yc   = 0.97
        _line_hc    = 0.88 / _n_cap
        ax_lc.text(0.5, 1.01, "Upside (%)", ha='center', va='bottom',
                   fontsize=8, fontweight='bold', color='#1C2B4A',
                   transform=ax_lc.transAxes)
        for _rank, _ki in enumerate(_sorted_cap):
            _col_d   = _cap_cols[_ki]
            _row_y   = _panel_yc - _rank * _line_hc
            ax_lc.plot(0.06, _row_y, 'o', color=_col_d, markersize=8,
                       transform=ax_lc.transAxes, clip_on=False)
            ax_lc.text(0.17, _row_y, labels_cap[_ki], ha='left', va='center',
                       fontsize=7.5, fontweight='bold', color='#1a1a2e',
                       transform=ax_lc.transAxes)
            ax_lc.text(0.98, _row_y, f"{y_cap[_ki]:+.1f}%", ha='right', va='center',
                       fontsize=7.5, color=_col_d, fontweight='bold',
                       transform=ax_lc.transAxes)

        st.pyplot(fig_cap)
        plt.close(fig_cap)


    # =====================================================
    # 10. CALENDAR RETURNS & ALPHA
    # =====================================================
    st.divider()
    cal_base = cleaned_df.set_index('Date')[manager_cols]
    cal_ret  = cal_base.groupby(cal_base.index.year).apply(
        lambda x: (np.prod(x + 1, axis=0) - 1) * 100).sort_index(ascending=False)
    st.write("**Calendar Returns (%)**")
    st.dataframe(style_df(cal_ret), width="stretch")

    bench_diff = st.selectbox("Select Alpha Benchmark", options=manager_cols, index=len(manager_cols)-1)
    cal_diff   = cal_ret.subtract(cal_ret[bench_diff], axis=0)
    st.write(f"**Calendar Difference (Alpha vs {bench_diff}) %**")
    st.dataframe(style_df(cal_diff), width="stretch")

    st.divider()
    alpha_fund  = st.selectbox("Select Fund for Alpha Matrix",      options=manager_cols, index=0)
    alpha_bench = st.selectbox("Select Benchmark for Alpha Matrix", options=manager_cols, index=len(manager_cols)-1)
    st.write(f"**Alpha Over Benchmark (Yearly View): {alpha_fund} vs {alpha_bench} (%)**")
######################################################################################################
    dates = cleaned_df["Date"]
    last_date = dates.max()
    latest_year = last_date.year

    f_series = cleaned_df.set_index("Date")[alpha_fund]
    b_series = cleaned_df.set_index("Date")[alpha_bench]

    years = sorted(dates.dt.year.unique())

    alpha_records = []

    for yr in years:

        if yr == latest_year:
            end_date = last_date
        else:
            end_date = pd.Timestamp(f"{yr}-12-31")

        row = {"Year": yr}

        for y in range(1, 21):

            months = y * 12

            # take last N months ending at end_date
            f_window = f_series.loc[:end_date].tail(months)
            b_window = b_series.loc[:end_date].tail(months)

            if len(f_window) == months and len(b_window) == months:

                f_ret = (np.prod(1 + f_window) ** (12 / months) - 1) * 100
                b_ret = (np.prod(1 + b_window) ** (12 / months) - 1) * 100

                row[f"{y}Y"] = f_ret - b_ret

            else:
                row[f"{y}Y"] = np.nan

        alpha_records.append(row)

    alpha_disp_year = pd.DataFrame(alpha_records).set_index("Year").sort_index(ascending=False)

    st.dataframe(style_df(alpha_disp_year), width="stretch", height=600)
#############################################################################################################

    # =====================================================
    # 11. CAPTURE TABLE
    # =====================================================
    st.divider()
    st.subheader("Upside / Downside Capture Table")
    cap_bench = st.selectbox("Benchmark for Capture Table", options=manager_cols,
                              index=len(manager_cols)-1, key="cap_tbl_bench")
    cap_horizons_lbl = [f"{y} Year" for y in [1, 3, 5, 8, 10]] + ["Since Inception"]
    cap_rows = []
    for mgr in manager_cols:
        row = {"Manager": mgr}
        for lbl_c in cap_horizons_lbl:
            yrs_c = int(lbl_c.split()[0]) if "Year" in lbl_c else None
            sm = cleaned_df[mgr].dropna(); sb = cleaned_df[cap_bench].dropna()
            if yrs_c:
                if len(sm) < yrs_c*12:
                    row[f"{lbl_c} Up"] = np.nan; row[f"{lbl_c} Dn"] = np.nan; continue
                sm = sm.tail(yrs_c*12); sb = sb.tail(yrs_c*12)
            u, d = get_cap(sm, sb)
            row[f"{lbl_c} Up"] = round(u, 2) if u is not None else np.nan
            row[f"{lbl_c} Dn"] = round(d, 2) if d is not None else np.nan
        cap_rows.append(row)
    cap_tbl = pd.DataFrame(cap_rows).set_index("Manager")

    # ── Win Rate (needed for Excel export) ──
    wr_bench = manager_cols[-1]
    wr_horizons = [("1 Year",1),("3 Year",3),("5 Year",5),("8 Year",8),("10 Year",10)]
    wr_rows = []
    for mgr in manager_cols:
        row_wr = {"Manager": mgr}
        for lbl_wr, yrs_wr in wr_horizons:
            sm_wr = cleaned_df[mgr].dropna(); sb_wr = cleaned_df[wr_bench].dropna()
            if yrs_wr and len(sm_wr) >= yrs_wr*12:
                sm_wr = sm_wr.tail(yrs_wr*12); sb_wr = sb_wr.tail(yrs_wr*12)
            combined_wr = pd.DataFrame({'m': sm_wr,'b': sb_wr}).dropna()
            row_wr[lbl_wr] = (combined_wr['m'] > combined_wr['b']).mean()*100 if len(combined_wr)>0 else np.nan
        wr_rows.append(row_wr)
    wr_df = pd.DataFrame(wr_rows).set_index("Manager")

    def style_capture(df_in):
        def color_cell(val):
            if pd.isna(val): return ''
            return 'color:green;font-weight:bold;text-align:center' if val >= 100 else 'color:red;text-align:center'
        return df_in.style.map(color_cell).format("{:.1f}", na_rep="").set_table_styles([
            {'selector': 'th.row_heading', 'props': [('font-weight','bold'),('color','black'),('text-align','left')]},
            {'selector': 'th.col_heading', 'props': [('font-weight','bold'),('color','black'),('text-align','center'),('font-size','11px')]},
        ])
    st.dataframe(style_capture(cap_tbl), width="stretch")

    # =====================================================
    # 13. MANAGER COMPARISON CARD
    # =====================================================
    st.divider()
    st.subheader("Manager Comparison")
    col_c1, col_c2 = st.columns(2)
    with col_c1: mgr_a = st.selectbox("Manager A", options=manager_cols, index=0, key="cmp_a")
    with col_c2: mgr_b = st.selectbox("Manager B", options=manager_cols, index=min(1,len(manager_cols)-1), key="cmp_b")
    cmp_horizon = st.selectbox("Comparison Horizon", options=h_labels, index=len(h_labels)-1, key="cmp_h")

    cmp_metrics = {
        "Annualized Return (%)":     ("{:.2f}%", True),
        "Annualized Volatility (%)": ("{:.2f}%", False),
        "Sharpe Ratio":              ("{:.2f}",  True),
        "Sortino Ratio":             ("{:.2f}",  True),
        "Max Drawdown (%)":          ("{:.2f}%", False),
        "Upward Deviation (%)":      ("{:.2f}%", True),
        "Downward Deviation (%)":    ("{:.2f}%", False),
    }

    st.markdown("""
    <style>
    .cmp-wrap{display:flex;gap:12px;margin-top:10px}
    .cmp-box{flex:1;border-radius:10px;padding:18px;background:#f8f9fa;border:1px solid #dee2e6}
    .cmp-title{font-size:18px;font-weight:800;color:#1a1a2e;margin-bottom:14px;border-bottom:3px solid #1a5276;padding-bottom:6px}
    .cmp-row{display:flex;justify-content:space-between;padding:6px 0;border-bottom:1px solid #eee;font-size:13px}
    .cmp-label{color:#555;font-weight:500}
    .cmp-val-good{color:#1e8449;font-weight:700}
    .cmp-val-bad{color:#922b21;font-weight:700}
    .cmp-val-neu{color:#333;font-weight:700}
    .cmp-winner{background:#eafaf1;border:2px solid #1e8449}
    </style>""", unsafe_allow_html=True)

    def get_cmp_val(mgr, metric):
        try: return float(results[metric].at[cmp_horizon, mgr])
        except: return np.nan

    def render_cmp_box(mgr, other, is_winner):
        box_class = "cmp-box cmp-winner" if is_winner else "cmp-box"
        rows_html = ""
        for metric, (fmt, higher_better) in cmp_metrics.items():
            v = get_cmp_val(mgr, metric); o = get_cmp_val(other, metric)
            if pd.isna(v): val_str, cls = "N/A", "cmp-val-neu"
            else:
                val_str = fmt.format(v)
                if pd.isna(o): cls = "cmp-val-neu"
                elif higher_better: cls = "cmp-val-good" if v > o else ("cmp-val-bad" if v < o else "cmp-val-neu")
                else:               cls = "cmp-val-good" if v < o else ("cmp-val-bad" if v > o else "cmp-val-neu")
            rows_html += f'<div class="cmp-row"><span class="cmp-label">{metric}</span><span class="{cls}">{val_str}</span></div>'
        badge = " 🏆" if is_winner else ""
        return f'<div class="{box_class}"><div class="cmp-title">{mgr}{badge}</div>{rows_html}</div>'

    sharpe_a = get_cmp_val(mgr_a, "Sharpe Ratio"); sharpe_b = get_cmp_val(mgr_b, "Sharpe Ratio")
    a_wins   = (not pd.isna(sharpe_a)) and (pd.isna(sharpe_b) or sharpe_a >= sharpe_b)
    st.markdown(f'<div class="cmp-wrap">{render_cmp_box(mgr_a,mgr_b,a_wins)}{render_cmp_box(mgr_b,mgr_a,not a_wins)}</div>',
                unsafe_allow_html=True)
    st.caption("🟢 = better value for that metric  |  🏆 = overall winner by Sharpe Ratio")

    # =====================================================
    # =====================================================
    # 14. DRAWDOWN TIMELINE — Professional Underwater Chart
    # =====================================================
    st.divider()
    st.subheader("Drawdown Timeline — Underwater Chart")
    dd_mgrs = st.multiselect("Managers for Drawdown Chart", options=manager_cols,
                               default=manager_cols[:4], key="dd_mgrs")
    if dd_mgrs:
        palette = ["#C0392B","#2471A3","#1E8449","#D35400","#6C3483","#117A65","#B7950B","#5D6D7E"]
        n_mgrs  = len(dd_mgrs)
        fig_dd, axes = plt.subplots(n_mgrs, 1, figsize=(15, 4.5*n_mgrs), sharex=False)
        if n_mgrs == 1: axes = [axes]
        fig_dd.patch.set_facecolor('#FAFBFC')

        for i, mgr in enumerate(dd_mgrs):
            s       = cleaned_df[mgr].fillna(0)
            dates   = cleaned_df['Date'].reset_index(drop=True)
            cum     = (1 + s.reset_index(drop=True)).cumprod()
            roll_mx = cum.cummax()
            dd_vals = (cum / roll_mx - 1) * 100

            ax  = axes[i]
            col = palette[i % len(palette)]

            ax.set_facecolor('#FFFFFF')
            # Gradient fill: underwater area
            ax.fill_between(dates, dd_vals, 0, where=(dd_vals < 0),
                            color=col, alpha=0.18, zorder=2, interpolate=True)
            ax.plot(dates, dd_vals, color=col, lw=1.6, zorder=3)
            ax.axhline(0, color='#1C2B4A', lw=0.9, zorder=4)

            # ── Year vertical lines + labels at top ──
            years = sorted(dates.dt.year.unique())
            y_min = dd_vals.min()
            for yr in years:
                yr_date = pd.Timestamp(f"{yr}-01-01")
                if dates.min() <= yr_date <= dates.max():
                    ax.axvline(yr_date, color='#DDDDDD', lw=0.7, ls='--', zorder=1)
                    ax.text(yr_date, 0.8, str(yr), fontsize=7, color='#888888',
                            ha='center', va='bottom', transform=ax.get_xaxis_transform(),
                            rotation=0, clip_on=True)

            # ── Max drawdown: find using positional index ──
            pos_min = int(dd_vals.values.argmin())   # positional index, always safe
            max_dd  = float(dd_vals.iloc[pos_min])
            max_date= dates.iloc[pos_min]

            # Recovery: first position after pos_min where dd >= -0.1%
            post_vals = dd_vals.iloc[pos_min:].reset_index(drop=True)
            rec_mask  = post_vals >= -0.1
            if rec_mask.any():
                rec_pos_local = int(rec_mask.values.argmax())
                rec_pos_abs   = pos_min + rec_pos_local
                rec_date_ts   = dates.iloc[rec_pos_abs]
            else:
                rec_pos_abs  = None
                rec_date_ts  = None

            # Drawdown start: last position before pos_min where dd >= -0.1%
            pre_vals   = dd_vals.iloc[:pos_min].reset_index(drop=True)
            start_mask = pre_vals >= -0.1
            if start_mask.any():
                start_pos_local = len(pre_vals) - 1 - int(start_mask.values[::-1].argmax())
                dd_start_ts     = dates.iloc[start_pos_local]
            else:
                dd_start_ts = dates.iloc[0]

            # ── Shade drawdown period ──
            if rec_date_ts is not None:
                ax.axvspan(dd_start_ts, rec_date_ts, alpha=0.07, color=col, zorder=1)
                dur_months = (rec_date_ts.year - dd_start_ts.year)*12 + (rec_date_ts.month - dd_start_ts.month)
                mid_date   = dd_start_ts + (rec_date_ts - dd_start_ts) / 2
                ax.text(mid_date, y_min * 0.55, f"{dur_months}m recovery",
                        fontsize=7.5, color='#555555', ha='center', va='center',
                        style='italic',
                        bbox=dict(boxstyle='round,pad=0.2', fc='white', ec='none', alpha=0.7))

            # ── Peak: last date before max-DD trough where cumulative == rolling max ──
            pre_cum   = cum.iloc[:pos_min+1]
            peak_iloc = int(pre_cum.values.argmax())
            peak_date = dates.iloc[peak_iloc]

            # ── Trough marker — red pill, always below the dot ──
            ax.scatter([max_date], [max_dd], color='#CC0000', s=80,
                       zorder=8, clip_on=False, edgecolors='white', linewidths=1.0)
            ax.annotate(
                f"Trough  {max_dd:.1f}%\n{max_date.strftime('%b %Y')}",
                xy=(max_date, max_dd),
                xytext=(0, -36), textcoords='offset points',
                fontsize=7.5, fontweight='bold', color='white', ha='center', va='top',
                arrowprops=dict(arrowstyle='->', color='#CC0000', lw=1.2,
                                connectionstyle='arc3,rad=0.0'),
                bbox=dict(boxstyle='round,pad=0.4', fc='#CC0000', ec='none', alpha=0.95),
                zorder=9
            )

            # ── Peak & Recovery: offset horizontally if they are close ──
            # Compute months apart — if < 12m apart, stagger them vertically
            _peak_top_off  = 22   # px above zero line
            _rec_top_off   = 22
            _peak_x_off    = 0    # horizontal pixel nudge
            _rec_x_off     = 0

            if rec_date_ts is not None:
                _months_apart = abs((rec_date_ts.year - peak_date.year)*12 +
                                    (rec_date_ts.month - peak_date.month))
                if _months_apart < 18:
                    # Stagger: peak stays low, recovery goes higher
                    _peak_top_off = 18
                    _rec_top_off  = 46
                    _peak_x_off   = -20
                    _rec_x_off    = 20

            # ── Peak marker — green pill above the zero line ──
            ax.scatter([peak_date], [0], color='#1E8449', s=80,
                       zorder=8, clip_on=False, edgecolors='white', linewidths=1.0)
            ax.annotate(
                f"Peak\n{peak_date.strftime('%b %Y')}",
                xy=(peak_date, 0),
                xytext=(_peak_x_off, _peak_top_off), textcoords='offset points',
                fontsize=7.5, fontweight='bold', color='white', ha='center', va='bottom',
                arrowprops=dict(arrowstyle='->', color='#1E8449', lw=1.0,
                                connectionstyle='arc3,rad=0.0'),
                bbox=dict(boxstyle='round,pad=0.4', fc='#1E8449', ec='none', alpha=0.95),
                zorder=9
            )

            # ── Recovery date marker — blue pill above zero line ──
            if rec_date_ts is not None:
                ax.scatter([rec_date_ts], [0], color='#1A5276', s=80,
                           zorder=8, clip_on=False, edgecolors='white', linewidths=1.0)
                ax.annotate(
                    f"Recovery\n{rec_date_ts.strftime('%b %Y')}",
                    xy=(rec_date_ts, 0),
                    xytext=(_rec_x_off, _rec_top_off), textcoords='offset points',
                    fontsize=7.5, fontweight='bold', color='white', ha='center', va='bottom',
                    arrowprops=dict(arrowstyle='->', color='#1A5276', lw=1.0,
                                    connectionstyle='arc3,rad=0.0'),
                    bbox=dict(boxstyle='round,pad=0.4', fc='#1A5276', ec='none', alpha=0.95),
                    zorder=9
                )
                ax.axvline(rec_date_ts, color=col, lw=0.9, ls=':', alpha=0.5, zorder=3)

            # ── Year axis on every panel ──
            ax.xaxis.set_major_locator(mdates.YearLocator())
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
            ax.tick_params(axis='x', labelsize=8, rotation=45)

            ax.set_ylabel("Drawdown (%)", fontsize=8, color='#444444')
            ax.set_title(f"  {mgr}", fontweight='bold', fontsize=11,
                         color='white', pad=0,
                         loc='left',
                         bbox=dict(boxstyle='square,pad=0.4', fc=col, ec='none'))
            ax.set_ylim(top=max(4, abs(y_min) * 0.18))   # enough headroom for Peak/Recovery labels
            ax.grid(True, axis='y', ls=':', alpha=0.35, zorder=0)
            ax.spines[['top','right']].set_visible(False)
            ax.spines['left'].set_color('#DDDDDD')
            ax.spines['bottom'].set_color('#DDDDDD')
            ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.0f}%"))
            ax.tick_params(labelsize=8, colors='#555555')



        fig_dd.suptitle("Drawdown — Underwater Analysis", fontsize=14,
                         fontweight='bold', color='#1C2B4A', y=1.005)
        plt.tight_layout(h_pad=1.8)
        st.pyplot(fig_dd)
        plt.close(fig_dd)

    # 16. AI STRATEGIC CONVERSATION
    # =====================================================
    st.divider()
    st.subheader("Jarir AI Strategic Advisor")

    full_results_text = ""
    for metric_name, df_res in results.items():
        full_results_text += f"\n--- {metric_name} ---\n{df_res.to_string()}\n"

    chat_context = f"""
    Portfolio Window: {start_label} to {end_label}
    Managers: {', '.join(manager_cols)}
    Alpha Check: {alpha_fund} vs {alpha_bench}
    [DATASET]\n{full_results_text}\n{alpha_disp_year.to_string()}
    """

    with st.expander("Activate AI Conversation", expanded=False):
        key_input = st.text_input("Enter OpenAI API Key", type="password")
        if st.button(" Generate insights"):
            if not key_input:
                st.error("Please provide an API key.")
            else:
                with st.spinner("Generating Insights..."):
                    try:
                        client = openai.OpenAI(api_key=key_input)
                        sys_role = "You are acting as a senior investment analyst performing a comprehensive evaluation of fund manager performance for institutional portfolio decisions. I will provide manager-level performance data including returns, volatility, upside and downside deviations, Sharpe and Sortino ratios, drawdowns, capture ratios, calendar returns, alpha metrics, and risk-return visualizations. Your objective is to interpret the data the way an experienced investment analyst would—by identifying what is really driving performance, when it occurred, and whether it is repeatable. Examine whether returns are generated through genuine manager skill or through exposure to favorable market regimes, elevated risk-taking, or beta concentration. Assess return consistency across time by identifying periods of performance concentration, regime dependence, and month-year-specific inflection points such as market stress, drawdowns, recoveries, or rallies. Evaluate downside risk by analyzing drawdowns, downside deviation, recovery speed, and Sortino behavior to determine capital preservation capability. Analyze upside versus downside capture to understand payoff asymmetry and to distinguish convex return profiles from leveraged or directional exposure. Use the risk-return positioning to identify efficiency, dominance, and risk-adjusted attractiveness relative to peers. Leverage alpha and relative metrics to test persistence, robustness, and benchmark independence, flagging statistically fragile or benchmark-hugging strategies. Synthesize these findings into actionable conclusions by classifying managers as core, satellite, tactical, or unsuitable; identifying complementary pairings based on risk and asymmetry; highlighting red flags and monitoring triggers; and presenting insights in clear, decision-focused language suitable for investment committee review, explicitly referencing relevant time periods and market contexts rather than relying on generic performance summaries."
                        user_prompt = f"Using this EXACT DATA: {chat_context}\n\nTask: {sys_role}"
                        response = client.chat.completions.create(
                            model="gpt-4-turbo",
                            messages=[{"role":"system","content":sys_role},{"role":"user","content":user_prompt}])
                        st.session_state.insights = response.choices[0].message.content
                        st.session_state.chat_history = [{"role":"assistant","content":st.session_state.insights}]
                    except Exception as e: st.error(f"AI Error: {e}")

        if 'insights' in st.session_state:
            st.write(st.session_state.insights)
            st.divider()
            st.write("**Chat with Jarir AI to get more insights:**")
            if "chat_history" not in st.session_state: st.session_state.chat_history = []
            for msg in st.session_state.chat_history:
                with st.chat_message(msg["role"]): st.write(msg["content"])
            if user_query := st.chat_input("Ask about the data..."):
                st.session_state.chat_history.append({"role":"user","content":user_query})
                with st.chat_message("user"): st.write(user_query)
                with st.chat_message("assistant"):
                    placeholder = st.empty(); placeholder.markdown("🤓 Investment Analyst is thinking...")
                    try:
                        client = openai.OpenAI(api_key=key_input)
                        chat_resp = client.chat.completions.create(
                            model="gpt-4-turbo",
                            messages=[{"role":"system","content":f"Senior Analyst. Context: {chat_context}"},
                                      *st.session_state.chat_history])
                        reply = chat_resp.choices[0].message.content
                        placeholder.empty(); st.write(reply)
                        st.session_state.chat_history.append({"role":"assistant","content":reply})
                    except Exception as e: st.error(f"Chat Error: {e}")

    # =====================================================
    # 17. EXCEL EXPORT
    # =====================================================
    def generate_excel(res, mgrs, cal_ret_df, cal_diff_df, alpha_df, cap_tbl_df, wr_df_ex):
        wb = Workbook(); wb.remove(wb.active)
        hdr_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        hdr_fill = PatternFill('solid', start_color='1A5276')
        idx_font = Font(name='Arial', bold=True, size=10)
        num_font = Font(name='Arial', size=10)
        red_font = Font(name='Arial', size=10, color='FF0000')
        center   = Alignment(horizontal='center', vertical='center')
        left_al  = Alignment(horizontal='left',   vertical='center')
        thin     = Side(style='thin', color='CCCCCC')
        bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill = PatternFill('solid', start_color='EBF5FB')

        def write_sheet(sheet_name, df, fmt="{:.2f}%"):
            ws = wb.create_sheet(title=sheet_name[:31]); ws.freeze_panes = 'B2'
            ws.cell(1,1,"Horizon").font = hdr_font; ws.cell(1,1).fill = hdr_fill
            ws.cell(1,1).alignment = center; ws.cell(1,1).border = bdr
            for ci, col in enumerate(df.columns, 2):
                c = ws.cell(1,ci,str(col))
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = center; c.border = bdr
                ws.column_dimensions[get_column_letter(ci)].width = max(14, len(str(col))+4)
            ws.column_dimensions['A'].width = 18; ws.row_dimensions[1].height = 22
            for ri, (idx, row) in enumerate(df.iterrows(), 2):
                ic = ws.cell(ri,1,str(idx)); ic.font = idx_font; ic.alignment = left_al; ic.border = bdr
                if ri%2==0: ic.fill = alt_fill
                for ci, val in enumerate(row, 2):
                    cell = ws.cell(ri,ci); cell.border = bdr; cell.alignment = center
                    if ri%2==0: cell.fill = alt_fill
                    try:
                        v = float(val)
                        cell.value = v/100 if "%" in fmt else v
                        cell.number_format = '0.00%' if "%" in fmt else '0.00'
                        cell.font = red_font if v < 0 else num_font
                    except: cell.value = ""; cell.font = num_font

        for metric in res:
            fmt = "{:.2f}%" if "Ratio" not in metric else "{:.2f}"
            write_sheet(metric[:31], res[metric].apply(pd.to_numeric, errors='coerce'), fmt)
        write_sheet("Calendar Returns",    cal_ret_df.apply(pd.to_numeric, errors='coerce'))
        write_sheet("Calendar Difference", cal_diff_df.apply(pd.to_numeric, errors='coerce'))
        write_sheet("Alpha Matrix",        alpha_df.apply(pd.to_numeric, errors='coerce'))

        # Capture table
        ws_cap = wb.create_sheet("Capture Table"); ws_cap.freeze_panes = 'B2'
        ws_cap.cell(1,1,"Manager").font = hdr_font; ws_cap.cell(1,1).fill = hdr_fill
        ws_cap.cell(1,1).alignment = center; ws_cap.cell(1,1).border = bdr
        ws_cap.column_dimensions['A'].width = 22
        for ci, col in enumerate(cap_tbl_df.columns, 2):
            c = ws_cap.cell(1,ci,str(col)); c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = bdr
            ws_cap.column_dimensions[get_column_letter(ci)].width = 12
        for ri, (idx, row) in enumerate(cap_tbl_df.iterrows(), 2):
            ws_cap.cell(ri,1,str(idx)).font = idx_font; ws_cap.cell(ri,1).alignment = left_al; ws_cap.cell(ri,1).border = bdr
            for ci, val in enumerate(row, 2):
                cell = ws_cap.cell(ri,ci); cell.border = bdr; cell.alignment = center
                try:
                    v = float(val); cell.value = v; cell.number_format = '0.0'
                    cell.font = Font(name='Arial',size=10,color='1E8449' if v>=100 else 'FF0000',bold=(v>=100))
                except: cell.value = ""; cell.font = num_font

        # Win Rate
        ws_wr = wb.create_sheet("Win Rate vs Benchmark"); ws_wr.freeze_panes = 'B2'
        ws_wr.cell(1,1,"Manager").font = hdr_font; ws_wr.cell(1,1).fill = hdr_fill
        ws_wr.cell(1,1).alignment = center; ws_wr.cell(1,1).border = bdr
        ws_wr.column_dimensions['A'].width = 22
        for ci, col in enumerate(wr_df_ex.columns, 2):
            c = ws_wr.cell(1,ci,str(col)); c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = bdr
            ws_wr.column_dimensions[get_column_letter(ci)].width = 10
        for ri, (idx, row) in enumerate(wr_df_ex.iterrows(), 2):
            ws_wr.cell(ri,1,str(idx)).font = idx_font; ws_wr.cell(ri,1).alignment = left_al; ws_wr.cell(ri,1).border = bdr
            for ci, val in enumerate(row, 2):
                cell = ws_wr.cell(ri,ci); cell.border = bdr; cell.alignment = center
                try:
                    v = float(val); cell.value = v/100; cell.number_format = '0.0%'
                    if v>=55:   cell.font = Font(name='Arial',size=10,color='1E8449',bold=True)
                    elif v<=45: cell.font = Font(name='Arial',size=10,color='FF0000')
                    else:       cell.font = Font(name='Arial',size=10,color='E67E22')
                except: cell.value = ""; cell.font = num_font

        buf = io.BytesIO(); wb.save(buf); buf.seek(0); return buf

    # =====================================================
    # 18. PRETTY PDF EXPORT
    # =====================================================

    # ─────────────────────────────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────────────────────────────
    def _fmt_is_pct(metric_name): return metric_name.endswith("(%)")

    def _format_val(v, is_pct=False):
        if pd.isna(v) or v == "": return ""
        try:
            f = float(v)
            return f"{f:,.2f}" + ("%" if is_pct else "")
        except: return str(v)

    def _is_negative(v):
        try: return float(v) < 0
        except: return False

    def _wrap_label(name, max_word_len=14):
        name = re.sub(r"[/\-]", " / ", name)
        def chunk(tok): return tok if len(tok)<=max_word_len else "\u200b".join([tok[i:i+max_word_len] for i in range(0,len(tok),max_word_len)])
        return " ".join([chunk(t) for t in name.split()])

    def _fig_to_rl_image(fig, max_w, max_h, dpi=150):
        buf = io.BytesIO(); fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight"); buf.seek(0)
        img = RLImage(buf); w, h = img.imageWidth, img.imageHeight
        scale = min(max_w/float(w), max_h/float(h), 1.0)
        img.drawWidth = w*scale; img.drawHeight = h*scale; return img

    # ─────────────────────────────────────────────────────────────────────
    # MAIN PDF GENERATOR
    # ─────────────────────────────────────────────────────────────────────
    def generate_pdf_pretty(results_dict, mgrs, rfr_name, start_l, end_l,
                             metrics_order, fig_rr_=None, fig_ud_=None, viz_h_="",alpha_bench=None, alpha_fund=None,
                             max_cols_per_page=20):

        buffer  = io.BytesIO()
        doc     = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                    leftMargin=20, rightMargin=20,
                                    topMargin=30, bottomMargin=28)
        page_w, page_h = landscape(A4)
        avail_w = page_w - doc.leftMargin - doc.rightMargin   # ~801 pt
        avail_h = page_h - doc.topMargin  - doc.bottomMargin
        styles  = getSampleStyleSheet()
        elems   = []

        # ── colour palette ──────────────────────────────────────────────
        C_NAVY   = colors.HexColor("#1C2B4A")
        C_LIGHT  = colors.HexColor("#EEF2F7")
        C_STRIPE = colors.HexColor("#F6F8FB")
        C_GRID   = colors.HexColor("#CFD8DC")
        C_RED    = colors.HexColor("#C0392B")
        C_GREEN  = colors.HexColor("#1E8449")
        C_WHITE  = colors.white

        # ── paragraph styles ────────────────────────────────────────────
        HDR  = ParagraphStyle("H", fontName="Helvetica-Bold", fontSize=7.5,
                               leading=9, alignment=TA_CENTER, textColor=C_WHITE)
        IDX  = ParagraphStyle("I", fontName="Helvetica-Bold", fontSize=7.5,
                               leading=9, alignment=TA_LEFT,   textColor=C_NAVY)
        CEL  = ParagraphStyle("C", fontName="Helvetica",      fontSize=7.5,
                               leading=9, alignment=TA_RIGHT)
        CEL_R= ParagraphStyle("CR",fontName="Helvetica",      fontSize=7.5,
                               leading=9, alignment=TA_RIGHT,  textColor=C_RED)
        SEC  = ParagraphStyle("S", fontName="Helvetica-Bold", fontSize=13,
                               leading=16, textColor=C_NAVY, spaceBefore=6, spaceAfter=3)
        SUB  = ParagraphStyle("SB",fontName="Helvetica",      fontSize=8,
                               leading=10, textColor=colors.HexColor("#555555"), spaceAfter=4)
        CAP  = ParagraphStyle("CP",fontName="Helvetica-Oblique", fontSize=7,
                               leading=8.5, textColor=colors.HexColor("#777777"), spaceAfter=2)

        # ── column width calculator ──────────────────────────────────────
        def col_widths_for(n_data_cols, idx_w=90):
            """Fit n data columns into available width."""
            rest = avail_w - idx_w
            each = max(44, min(rest / n_data_cols, 110))
            # re-check total fits
            total = idx_w + each * n_data_cols
            if total > avail_w:
                each = (avail_w - idx_w) / n_data_cols
            return [idx_w] + [each] * n_data_cols

        # ── base table style ─────────────────────────────────────────────
        def base_ts(nrows):
            ts = TableStyle([
                ("BACKGROUND", (0,0), (-1,0), C_NAVY),
                ("TEXTCOLOR", (0,0), (-1,0), C_WHITE),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 7.5),
                ("ALIGN", (0,0), (-1,0), "CENTER"),

                ("BOX", (0,0), (-1,-1), 1, colors.black),
                ("INNERGRID", (0,0), (-1,-1), 0.5, colors.black),

                ("LEFTPADDING", (0,0), (-1,-1), 4),
                ("RIGHTPADDING", (0,0), (-1,-1), 4),
                ("TOPPADDING", (0,0), (-1,-1), 3),
                ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ])
            # alternating rows
            for r in range(1, nrows):
                if r % 2 == 0:
                    ts.add("BACKGROUND", (0, r), (-1, r), C_STRIPE)
            return ts

        # ── metric table builder (handles chunking for >10 managers) ─────
        def add_metric_table(title, subtitle, df_in, is_pct=True, chunk_size=10):
            df_in = df_in.apply(pd.to_numeric, errors='coerce').round(2)
            n_mgrs = df_in.shape[1]
            chunks = max(1, math.ceil(n_mgrs / chunk_size))
            for ci in range(chunks):
                sc = ci * chunk_size
                ec = min(sc + chunk_size, n_mgrs)
                sub = df_in.iloc[:, sc:ec]
                mgr_names = list(sub.columns)

                elems.append(Paragraph(title, SEC))
                chunk_note = f" (Part {ci+1}/{chunks})" if chunks > 1 else ""
                elems.append(Paragraph(
                    f"<b>Managers:</b> {', '.join(mgr_names)}{chunk_note}  |  "
                    f"<b>Period:</b> {start_l} – {end_l}  |  <b>RFR:</b> {rfr_name}",
                    SUB))

                hdr_row = [Paragraph("Horizon", HDR)] + \
                          [Paragraph(_wrap_label(c), HDR) for c in mgr_names]
                rows = [hdr_row]
                for horizon, row in sub.iterrows():
                    r = [Paragraph(str(horizon), IDX)]
                    for v in row.values:
                        if pd.isna(v) or v == "":
                            r.append(Paragraph("—", CEL))
                        else:
                            txt = f"{v:,.2f}%" if is_pct else f"{v:,.2f}"
                            r.append(Paragraph(txt, CEL_R if v < 0 else CEL))
                    rows.append(r)

                cw  = col_widths_for(len(mgr_names))
                tbl = LongTable(rows, repeatRows=1, colWidths=cw, hAlign="LEFT")
                ts  = base_ts(len(rows))
                tbl.setStyle(ts)
                elems.append(tbl)
                if ci < chunks - 1:
                    elems.append(Spacer(1, 10))
                else:
                    elems.append(PageBreak())

        # ── simple wide table (calendar / monthly / correlation) ─────────
        def add_wide_table(title, df_in, is_pct=True, transpose=False, caption_txt="", chunk_size=10):
            if transpose:
                df_in = df_in.T
            elems.append(Paragraph(title, SEC))
            if caption_txt:
                elems.append(Paragraph(caption_txt, CAP))

            n_cols = df_in.shape[1]
            chunks = max(1, math.ceil(n_cols / chunk_size))
            for ci in range(chunks):
                sc = ci * chunk_size
                ec = min(sc + chunk_size, n_cols)
                sub = df_in.iloc[:, sc:ec]

                hdr_row = [Paragraph("", HDR)] + \
                          [Paragraph(_wrap_label(str(c)), HDR) for c in sub.columns]
                rows = [hdr_row]
                for idx_, row_ in sub.iterrows():
                    r = [Paragraph(str(idx_), IDX)]
                    for v in row_.values:
                        if pd.isna(v) or v == "":
                            r.append(Paragraph("—", CEL))
                        else:
                            try:
                                fv = float(v)
                                txt = f"{fv:,.2f}%" if is_pct else f"{fv:,.4f}"
                                r.append(Paragraph(txt, CEL_R if fv < 0 else CEL))
                            except:
                                r.append(Paragraph(str(v), CEL))
                    rows.append(r)

                cw  = col_widths_for(len(sub.columns), idx_w=70)
                tbl = LongTable(rows, repeatRows=1, colWidths=cw, hAlign="LEFT")
                ts  = base_ts(len(rows))
                tbl.setStyle(ts)
                elems.append(tbl)
                if ci < chunks - 1:
                    elems.append(Spacer(1, 8))

            elems.append(PageBreak())

        # ── scatter builder for PDF ──────────────────────────────────────
        def make_scatter_fig(x_vals, y_vals, labels, dot_color, xlabel, ylabel,
                              title, reflines=None):
            fig, ax = plt.subplots(figsize=(7, 4.2))
            ax.scatter(x_vals, y_vals, color=dot_color, s=70, zorder=5)
            x_arr = np.array(x_vals, dtype=float)
            y_arr = np.array(y_vals, dtype=float)
            n = len(labels)
            if n:
                xr = max(x_arr.max()-x_arr.min(), 1e-6)
                yr = max(y_arr.max()-y_arr.min(), 1e-6)
                lx = x_arr + xr*0.04; ly = y_arr + yr*0.04
                for _ in range(200):
                    for i in range(n):
                        fx, fy = 0.0, 0.0
                        for j in range(n):
                            if i==j: continue
                            dx=(lx[i]-lx[j])/xr; dy=(ly[i]-ly[j])/yr
                            d2=dx*dx+dy*dy+1e-9
                            if d2<0.08: fx+=dx/d2; fy+=dy/d2
                        for j in range(n):
                            dx=(lx[i]-x_arr[j])/xr; dy=(ly[i]-y_arr[j])/yr
                            d2=dx*dx+dy*dy+1e-9
                            if d2<0.06: fx+=0.5*dx/d2; fy+=0.5*dy/d2
                        lx[i]+=fx*0.0008*xr; ly[i]+=fy*0.0008*yr
                for i, lbl in enumerate(labels):
                    ax.annotate(lbl, xy=(x_arr[i],y_arr[i]), xytext=(lx[i],ly[i]),
                                fontsize=7, fontweight='bold', color='#1a1a2e',
                                bbox=dict(boxstyle='round,pad=0.25',fc='white',ec='#ccc',alpha=0.85,lw=0.5),
                                arrowprops=dict(arrowstyle='-',color='#aaa',lw=0.7), zorder=10)
            ax.set_xlabel(xlabel, fontsize=8); ax.set_ylabel(ylabel, fontsize=8)
            ax.set_title(title, fontsize=9, fontweight='bold', color='#1C2B4A')
            ax.grid(True, ls=':', alpha=0.5)
            if reflines:
                for axis, val, lw in reflines:
                    if axis=='h': ax.axhline(val, color='#333', lw=lw)
                    if axis=='v': ax.axvline(val, color='#333', lw=lw)
            plt.tight_layout()
            return fig

        # ── page footer ──────────────────────────────────────────────────
        def add_footer(canvas, doc_):
            canvas.saveState()
            w = doc_.pagesize[0]
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(colors.HexColor("#888888"))
            canvas.drawString(doc_.leftMargin, 14,
                f"Jarir Investments — Confidential  |  {start_l} to {end_l}")
            canvas.drawRightString(w - doc_.rightMargin, 14, f"Page {doc_.page}")
            canvas.restoreState()

        # ══════════════════════════════════════════════════════════════════
        # COVER PAGE
        # ══════════════════════════════════════════════════════════════════
        now_str = dt.datetime.now().strftime("%d %B %Y, %H:%M")
        cover_title = ParagraphStyle("CT", fontName="Helvetica-Bold", fontSize=22,
                                      leading=26, textColor=C_NAVY, alignment=TA_CENTER)
        cover_sub   = ParagraphStyle("CS", fontName="Helvetica", fontSize=11,
                                      leading=14, textColor=colors.HexColor("#444444"), alignment=TA_CENTER)
        cover_info  = ParagraphStyle("CI", fontName="Helvetica", fontSize=9,
                                      leading=12, textColor=colors.HexColor("#666666"), alignment=TA_CENTER)

        elems.append(Spacer(1, 80))
        elems.append(Paragraph("Jarir Investments", cover_title))
        elems.append(Spacer(1, 8))
        elems.append(Paragraph(f" {alpha_fund} Quantitative Performance Report", cover_sub))
        elems.append(Spacer(1, 20))
        elems.append(Paragraph(f"Analysis Period:  <b>{start_l}</b>  to  <b>{end_l}</b>", cover_sub))
        elems.append(Spacer(1, 10))
        elems.append(Paragraph(
            f"Managers ({len(mgrs)}):  {', '.join(mgrs)}", cover_info))
        elems.append(Spacer(1, 6))
        elems.append(Paragraph(f"Risk-Free Rate:  {rfr_name}", cover_info))
        elems.append(Spacer(1, 6))
        elems.append(Paragraph(f"Generated:  {now_str}", cover_info))
        elems.append(PageBreak())

        # ══════════════════════════════════════════════════════════════════
        # SECTION 1 — PERFORMANCE METRICS (7 tables)
        # ══════════════════════════════════════════════════════════════════
        ordered_metrics = [
            ("Annualized Return (%)",    True),
            ("Annualized Volatility (%)",True),
            ("Sharpe Ratio",             False),
            ("Sortino Ratio",            False),
            ("Downward Deviation (%)",   True),
            ("Upward Deviation (%)",     True),
            ("Max Drawdown (%)",         True),
        ]
        for metric_name, is_pct in ordered_metrics:
            if metric_name in results_dict:
                add_metric_table(
                    title    = metric_name,
                    subtitle = "",
                    df_in    = results_dict[metric_name][mgrs],
                    is_pct   = is_pct,
                    chunk_size = max_cols_per_page
                )

        # ══════════════════════════════════════════════════════════════════
        # SECTION 2 — CALENDAR RETURNS & DIFFERENTIAL
        # ══════════════════════════════════════════════════════════════════
        add_wide_table(
            "Calendar Year Returns (%)",
            cal_ret[[c for c in mgrs if c in cal_ret.columns]].apply(pd.to_numeric, errors='coerce').round(2),
            is_pct=True, chunk_size=max_cols_per_page,
            caption_txt="Full-year compounded returns for each calendar year."
        )
        add_wide_table(
            f"Calendar Differences",
            cal_diff[[c for c in mgrs if c in cal_diff.columns]].apply(pd.to_numeric, errors='coerce').round(2),
            is_pct=True, chunk_size=max_cols_per_page
        )

        # ══════════════════════════════════════════════════════════════════
        # SECTION 3 — MONTHLY RETURNS (trailing 36 months)
        # ══════════════════════════════════════════════════════════════════
        m36 = cleaned_df.set_index('Date')[mgrs].tail(36).iloc[::-1].copy()
        m36.index = m36.index.strftime('%b-%Y')
        add_wide_table(
            "Monthly Returns — Trailing 36 Months (%)",
            (m36 * 100).apply(pd.to_numeric, errors='coerce').round(2),
            is_pct=True, chunk_size=max_cols_per_page,
            caption_txt="Most recent 36 monthly returns, latest first."
        )

        # ══════════════════════════════════════════════════════════════════
        # SECTION 4 — CORRELATION MATRIX
        # ══════════════════════════════════════════════════════════════════
        corr_df = cleaned_df[mgrs].corr().round(4)
        elems.append(Paragraph("Correlation Matrix (Since Inception)", SEC))
        elems.append(Paragraph(
            "Pairwise Pearson correlation of monthly returns. Values closer to 1 indicate higher co-movement.",
            CAP))

        n_c = len(mgrs)
        chunks_c = max(1, math.ceil(n_c / max_cols_per_page))
        for ci in range(chunks_c):
            sc = ci * max_cols_per_page; ec = min(sc + max_cols_per_page, n_c)
            sub_c = corr_df.iloc[:, sc:ec]
            hdr_row = [Paragraph("", HDR)] + [Paragraph(_wrap_label(c), HDR) for c in sub_c.columns]
            rows = [hdr_row]
            for idx_, row_ in sub_c.iterrows():
                r = [Paragraph(str(idx_), IDX)]
                for v in row_.values:
                    if pd.isna(v): r.append(Paragraph("—", CEL))
                    else:
                        txt = f"{v:.4f}"
                        sty = CEL_R if v < 0 else CEL
                        r.append(Paragraph(txt, sty))
                rows.append(r)
            cw  = col_widths_for(len(sub_c.columns), idx_w=90)
            tbl = LongTable(rows, repeatRows=1, colWidths=cw, hAlign="LEFT")
            ts  = base_ts(len(rows))
            tbl.setStyle(ts)
            elems.append(tbl)
            if ci < chunks_c - 1:
                elems.append(Spacer(1, 8))
        elems.append(PageBreak())

        # ══════════════════════════════════════════════════════════════════
        # SECTION 5 — ALPHA OVER BENCHMARK
        # ══════════════════════════════════════════════════════════════════
        add_wide_table(
            f"Alpha Over Benchmark for {alpha_fund}",
            alpha_disp_year.apply(pd.to_numeric, errors='coerce').round(2),
            is_pct=True, chunk_size=max_cols_per_page
        
        )

        # ══════════════════════════════════════════════════════════════════
        # ══════════════════════════════════════════════════════════════════
        # ══════════════════════════════════════════════════════════════════
        # SECTION 6 — RISK–RETURN SCATTER (3Y, 5Y, 8Y) — one per page, full size
        # ══════════════════════════════════════════════════════════════════

        def _rr_fig_full(hz_label, dot_color="#0d6e8a"):
            if hz_label not in results_dict["Annualized Return (%)"].index:
                return None
            ret_s = pd.to_numeric(results_dict["Annualized Return (%)"].loc[hz_label],    errors="coerce")
            vol_s = pd.to_numeric(results_dict["Annualized Volatility (%)"].loc[hz_label], errors="coerce")
            rr = pd.DataFrame({"ret": ret_s, "vol": vol_s}).dropna()
            rr = rr[rr.index.isin(mgrs)]
            if rr.empty: return None

            fig, ax = plt.subplots(figsize=(12, 7))
            fig.patch.set_facecolor("#F9FAFB")
            ax.set_facecolor("#FAFBFC")

            # Colour dots by return level — viridis gradient
            norm = plt.Normalize(rr["ret"].min(), rr["ret"].max())
            cmap = plt.cm.RdYlGn
            sc = ax.scatter(rr["vol"], rr["ret"],
                            c=rr["ret"], cmap=cmap, norm=norm,
                            s=120, zorder=5, edgecolors="white", linewidths=0.8)
            cbar = fig.colorbar(sc, ax=ax, pad=0.02, fraction=0.03)
            cbar.set_label("Return (%)", fontsize=9, color="#555")
            cbar.ax.tick_params(labelsize=8)

            # Repulsion-based labels
            x_arr = np.array(rr["vol"].tolist(), dtype=float)
            y_arr = np.array(rr["ret"].tolist(), dtype=float)
            labels_rr = rr.index.tolist()
            n = len(labels_rr)
            xr = max(x_arr.max()-x_arr.min(), 1e-6)
            yr = max(y_arr.max()-y_arr.min(), 1e-6)
            lx = x_arr + xr*0.03; ly = y_arr + yr*0.04
            for _ in range(300):
                for ii in range(n):
                    fx, fy = 0.0, 0.0
                    for jj in range(n):
                        if ii==jj: continue
                        dx=(lx[ii]-lx[jj])/xr; dy=(ly[ii]-ly[jj])/yr
                        d2=dx*dx+dy*dy+1e-9
                        if d2<0.08: fx+=dx/d2; fy+=dy/d2
                    for jj in range(n):
                        dx=(lx[ii]-x_arr[jj])/xr; dy=(ly[ii]-y_arr[jj])/yr
                        d2=dx*dx+dy*dy+1e-9
                        if d2<0.06: fx+=0.5*dx/d2; fy+=0.5*dy/d2
                    lx[ii]+=fx*0.0007*xr; ly[ii]+=fy*0.0007*yr
            for ii, lbl in enumerate(labels_rr):
                ax.annotate(lbl,
                    xy=(x_arr[ii], y_arr[ii]), xytext=(lx[ii], ly[ii]),
                    fontsize=9, fontweight="bold", color="#1a1a2e",
                    bbox=dict(boxstyle="round,pad=0.35", fc="white", ec="#cccccc", alpha=0.9, lw=0.7),
                    arrowprops=dict(arrowstyle="-", color="#aaaaaa", lw=0.9), zorder=10)

            ax.set_xlabel("Annualized Volatility (%)", fontsize=11)
            ax.set_ylabel("Annualized Return (%)",     fontsize=11)
            ax.set_title(f"Risk–Return Analysis — {hz_label}",
                         fontsize=14, fontweight="bold", color="#1C2B4A", pad=12)
            ax.grid(True, ls=":", alpha=0.45, color="#CCCCCC")
            ax.spines[["top","right"]].set_visible(False)
            ax.spines[["left","bottom"]].set_color("#CCCCCC")
            ax.tick_params(labelsize=9)
            plt.tight_layout(pad=1.5)
            return fig

        for hz_rr in ["3 Year", "5 Year", "8 Year"]:
            f_rr = _rr_fig_full(hz_rr)
            if f_rr:
                elems.append(Paragraph(f"Risk–Return Analysis — {hz_rr}", SEC))
                elems.append(Paragraph(
                    f"Period: {start_l} – {end_l}  |  Each dot = one manager  |  "
                    f"X-axis: Annualized Volatility (%)  |  Y-axis: Annualized Return (%)  |  "
                    f"Colour gradient: Return level", CAP))
                elems.append(Spacer(1, 6))
                elems.append(_fig_to_rl_image(f_rr, avail_w, avail_h * 0.82))
                plt.close(f_rr)
                elems.append(PageBreak())

        # SECTION 7 — UPSIDE/DOWNSIDE CAPTURE PLOTS (3Y, 5Y, 8Y) + TABLE
        # ══════════════════════════════════════════════════════════════════
        elems.append(Paragraph("Upside / Downside Capture Analysis", SEC))
        elems.append(Paragraph(
            "X-axis: Downside Capture (lower = better)  |  Y-axis: Upside Capture (higher = better)  |  "
            "Dashed lines at 100% = benchmark baseline.",
            CAP))

        cap_bench_pdf = alpha_bench

        def _cap_fig(hz_lbl, yrs_c):
            rows = []
            for m in mgrs:
                sm = cleaned_df[m].dropna()
                sb = cleaned_df[cap_bench_pdf].dropna()
                if len(sm) < yrs_c * 12: continue
                sm = sm.tail(yrs_c * 12); sb = sb.tail(yrs_c * 12)
                u, d = get_cap(sm, sb)
                if pd.notna(u) and pd.notna(d):
                    rows.append({"Manager": m, "Up": u, "Dn": d})
            if not rows: return None, None
            cap_df = pd.DataFrame(rows).set_index("Manager")
            # fig = make_scatter_fig(
            #     cap_df["Dn"].tolist(), cap_df["Up"].tolist(), cap_df.index.tolist(),
            #     "#8e1a0e", "Downside Capture (%)", "Upside Capture (%)",
            #     f"Capture Ratio — {hz_lbl}  (Benchmark: {cap_bench_pdf})",
            #     reflines=[("h", 100, 0.8), ("v", 100, 0.8)]
            # )
            fig, ax = plt.subplots(figsize=(11,5.5))
            fig.patch.set_facecolor('#F9FAFB')
            ax.set_facecolor('#FAFBFC')

            x_cap = cap_df["Dn"].values
            y_cap = cap_df["Up"].values
            labels_cap = cap_df.index.tolist()

            norm = plt.Normalize(y_cap.min(), y_cap.max())
            cmap = plt.cm.RdYlGn

            sc = ax.scatter(
                x_cap,
                y_cap,
                c=y_cap,
                cmap=cmap,
                norm=norm,
                s=120,
                edgecolors='white',
                linewidths=0.8,
                zorder=5
            )

            # colorbar
            cbar = fig.colorbar(sc, ax=ax, pad=0.02)
            cbar.set_label("Upside (%)", fontsize=8)

            # labels
            for i, label in enumerate(labels_cap):
                ax.annotate(
                    label,
                    xy=(x_cap[i], y_cap[i]),
                    xytext=(x_cap[i]+1, y_cap[i]+1),
                    fontsize=8,
                    fontweight='bold',
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#CCCCCC", alpha=0.9),
                    arrowprops=dict(arrowstyle="-", color="#999999", lw=0.8)
                )

            # reference lines
            ax.axhline(100, color='#444444', lw=0.9, ls='--')
            ax.axvline(100, color='#444444', lw=0.9, ls='--')

            ax.set_xlabel("Downside Capture (%)")
            ax.set_ylabel("Upside Capture (%)")
            ax.set_title(f"Capture Ratio — {hz_lbl} (Benchmark: {cap_bench_pdf})", fontweight="bold")

            ax.grid(True, ls=':', alpha=0.4)
            ax.spines[['top','right']].set_visible(False)
            return fig, cap_df

        f3y, cd3 = _cap_fig("3 Year", 3)
        f5y, cd5 = _cap_fig("5 Year", 5)
        f8y, cd8 = _cap_fig("8 Year", 8)

        # Each capture horizon on its own page
        for lbl_cp, fig_cp in [("3 Year", f3y), ("5 Year", f5y), ("8 Year", f8y)]:
            if fig_cp is None: continue
            elems.append(Spacer(1, avail_h * 0.08))
            elems.append(_fig_to_rl_image(fig_cp, avail_w, avail_h * 0.70))
            plt.close(fig_cp)
            elems.append(PageBreak())

        # Capture summary table (Up/Dn columns for 3Y, 5Y, 8Y)
        elems.append(Paragraph("Upside / Downside Capture Summary Table", SEC))
        elems.append(Paragraph(
            f"Benchmark: {cap_bench_pdf}.  Values > 100 on Upside = outperforms benchmark in up-markets.  "
            f"Values < 100 on Downside = loses less than benchmark in down-markets.",
            CAP))

        cap_col_labels = ["3Y Up", "3Y Dn", "5Y Up", "5Y Dn", "8Y Up", "8Y Dn"]
        cap_summary = pd.DataFrame(index=mgrs, columns=cap_col_labels)
        for m in mgrs:
            for cd, u_lbl, d_lbl in [(cd3, "3Y Up", "3Y Dn"), (cd5, "5Y Up", "5Y Dn"), (cd8, "8Y Up", "8Y Dn")]:
                if cd is not None and m in cd.index:
                    cap_summary.at[m, u_lbl] = cd.at[m, "Up"]
                    cap_summary.at[m, d_lbl] = cd.at[m, "Dn"]

        hdr_cap = [Paragraph("Manager", HDR)] + [Paragraph(c, HDR) for c in cap_col_labels]
        rows_cap = [hdr_cap]
        for mgr_c, row_c in cap_summary.iterrows():
            r = [Paragraph(str(mgr_c), IDX)]
            for col_c, v in row_c.items():
                if pd.isna(v) or v == "":
                    r.append(Paragraph("—", CEL))
                else:
                    try:
                        fv = float(v)
                        sty = CEL_R if fv < 100 and "Dn" in col_c else CEL
                        r.append(Paragraph(f"{fv:.1f}%", sty))
                    except:
                        r.append(Paragraph(str(v), CEL))
            rows_cap.append(r)

        cw_cap = col_widths_for(len(cap_col_labels), idx_w=110)
        tbl_cap = LongTable(rows_cap, repeatRows=1, colWidths=cw_cap, hAlign="LEFT")
        tbl_cap.setStyle(base_ts(len(rows_cap)))
        elems.append(tbl_cap)
        elems.append(PageBreak())

        # ── build ─────────────────────────────────────────────────────────
        doc.build(elems, onFirstPage=add_footer, onLaterPages=add_footer)
        buffer.seek(0); return buffer.read()


with tab2:
    import plotly.express as px
    import plotly.graph_objects as go

    st.subheader("Institutional Deep-Dive")

    viz_mgrs = st.multiselect(
        "Select managers for visual comparison",
        options=manager_cols,
        default=manager_cols[:4],
        key="viz_mgrs_sv"
    )

    if len(viz_mgrs) > 0:
        col_v1, col_v2 = st.columns(2)

        with col_v1:
            st.write("**Annualized Return Trajectory**")
            plot_horizons = [f"{i} Year" for i in range(1, 11)]
            # Map to h_labels naming used in results
            traj_data = pd.DataFrame(index=plot_horizons, columns=viz_mgrs)
            for ph in plot_horizons:
                if ph in results["Annualized Return (%)"]:
                    pass
            # Build from results dict using h_labels names
            ph_labels = [lbl for lbl in h_labels if "Year" in lbl and int(lbl.split()[0]) <= 10]
            traj_df = results["Annualized Return (%)"].loc[ph_labels, viz_mgrs].apply(pd.to_numeric, errors="coerce")
            traj_df.index = [lbl.replace(" Year", "Y") for lbl in traj_df.index]
            viz_data_ann = traj_df.reset_index().rename(columns={"index": "Horizon"})
            fig_ann = px.line(viz_data_ann, x="Horizon", y=viz_mgrs, markers=True, template="plotly_white")
            st.plotly_chart(fig_ann, use_container_width=True)

        with col_v2:
            st.write("**Calendar Year Performance Heatmap**")
            cal_base_sv = cleaned_df.set_index("Date")[viz_mgrs]
            cal_years_sv = sorted(cleaned_df["Date"].dt.year.unique(), reverse=True)
            cal_years_sv = [y for y in cal_years_sv if y >= 2000]
            heatmap_df = pd.DataFrame(index=viz_mgrs, columns=[str(y) for y in cal_years_sv[:10]])
            for mgr_sv in viz_mgrs:
                s_sv = cal_base_sv[mgr_sv]
                for yr in cal_years_sv[:10]:
                    yd = s_sv[s_sv.index.year == yr]
                    if not yd.empty:
                        heatmap_df.at[mgr_sv, str(yr)] = round((np.prod(1 + yd) - 1) * 100, 2)
            heatmap_data = heatmap_df.astype(float)
            fig_heat = px.imshow(heatmap_data, color_continuous_scale="RdYlGn", aspect="auto", text_auto=".1f")
            fig_heat.update_xaxes(side="top")
            st.plotly_chart(fig_heat, use_container_width=True)

        st.divider()
        col_r1, col_r2 = st.columns(2)

        with col_r1:
            st.write("**Risk-Return Efficiency (Efficient Frontier)**")
            si_returns_sv = {}
            for mgr_sv in viz_mgrs:
                s_sv = cleaned_df[mgr_sv].dropna()
                si_returns_sv[mgr_sv] = ((np.prod(1 + s_sv) ** (12 / len(s_sv))) - 1) * 100 if len(s_sv) > 0 else np.nan
            vol_store_sv = {m: cleaned_df[m].dropna().std() * np.sqrt(12) * 100 for m in viz_mgrs}
            rr_df_sv = pd.DataFrame({
                "Manager":       viz_mgrs,
                "Return (%)":    [si_returns_sv[m] for m in viz_mgrs],
                "Volatility (%)": [vol_store_sv[m] for m in viz_mgrs]
            })
            fig_rr_sv = px.scatter(rr_df_sv, x="Volatility (%)", y="Return (%)",
                                    text="Manager", color="Return (%)", template="plotly_white")
            fig_rr_sv.update_traces(textposition="top center")
            st.plotly_chart(fig_rr_sv, use_container_width=True)

        with col_r2:
            st.write("**Growth of $100 (Cumulative Performance)**")
            returns_df_sv = cleaned_df.set_index("Date")[viz_mgrs]
            wealth_index_sv = (1 + returns_df_sv).cumprod() * 100
            start_row_sv = pd.DataFrame(
                100.0,
                index=[cleaned_df["Date"].min() - pd.Timedelta(days=1)],
                columns=viz_mgrs
            )
            wealth_index_sv = pd.concat([start_row_sv, wealth_index_sv]).sort_index()
            fig_growth_sv = px.line(wealth_index_sv, template="plotly_white",
                                     labels={"value": "Wealth ($)", "index": ""})
            st.plotly_chart(fig_growth_sv, use_container_width=True)

        st.divider()
        st.subheader("💰 Custom Portfolio Wealth Index")
        assign_weights = st.checkbox("Do you want to assign dollar weights to these managers?", key="assign_w_sv")

        if assign_weights:
            weight_inputs = st.columns(len(viz_mgrs))
            invested_amounts = {}
            for i, mgr_sv in enumerate(viz_mgrs):
                with weight_inputs[i]:
                    invested_amounts[mgr_sv] = st.number_input(
                        f"Assign to {mgr_sv} ($)",
                        min_value=0.0, max_value=100.0, value=0.0, step=5.0,
                        key=f"w_sv_{mgr_sv}"
                    )
            total_sum = sum(invested_amounts.values())
            st.write(f"**Total Assigned: ${total_sum:,.1f} / $100.0**")
            if total_sum > 100.1:
                st.error("Total allocated dollars exceed $100. Please adjust.")
            elif total_sum > 0:
                p_wealth = (1 + cleaned_df.set_index("Date")[viz_mgrs]).cumprod()
                for mgr_sv in viz_mgrs:
                    p_wealth[mgr_sv] = p_wealth[mgr_sv] * invested_amounts[mgr_sv]
                total_portfolio = p_wealth.sum(axis=1)
                fig_port = go.Figure()
                fig_port.add_trace(go.Scatter(
                    x=total_portfolio.index, y=total_portfolio,
                    name="CUSTOM PORTFOLIO", line=dict(color="black", width=3)
                ))
                for mgr_sv in viz_mgrs:
                    if invested_amounts[mgr_sv] > 0:
                        fig_port.add_trace(go.Scatter(
                            x=p_wealth.index, y=p_wealth[mgr_sv],
                            name=f"{mgr_sv} (${invested_amounts[mgr_sv]:.1f})",
                            line=dict(dash="dot", width=1)
                        ))
                fig_port.update_layout(
                    template="plotly_white",
                    title="Weighted Portfolio Performance",
                    yaxis_title="Value ($)"
                )
                st.plotly_chart(fig_port, use_container_width=True)

        # ─── Dynamic n-Year Rolling Alpha Analytics ───────────────────────
        st.divider()
        st.subheader("📊 Dynamic n-Year Rolling Alpha Suite")

        roll_n = st.number_input(
            "Select n-Year Window for Rolling Alpha",
            min_value=1, max_value=10, value=3, step=1, key="roll_n_sv2"
        )
        window_size = int(roll_n * 12)

        a_col1, a_col2 = st.columns(2)
        with a_col1:
            alpha_mgr = st.selectbox("Select Manager for Alpha",   options=manager_cols, key="a_mgr_sv2")
        with a_col2:
            alpha_bench = st.selectbox("Select Benchmark for Alpha", options=manager_cols,
                                        index=len(manager_cols)-1, key="a_bench_sv2")

        if alpha_mgr and alpha_bench:
            m_ret_sv = cleaned_df.set_index("Date")[alpha_mgr]
            b_ret_sv = cleaned_df.set_index("Date")[alpha_bench]

            def roll_ann_sv2(s, win):
                return s.rolling(window=win).apply(
                    lambda x: (np.prod(1 + x) ** (12 / win)) - 1 if len(x) == win else np.nan
                )

            m_ann_roll = roll_ann_sv2(m_ret_sv, window_size)
            b_ann_roll = roll_ann_sv2(b_ret_sv, window_size)
            roll_alpha = (m_ann_roll - b_ann_roll).dropna()

            if not roll_alpha.empty:
                dynamic_avg = roll_alpha.rolling(window=window_size).mean()
                dynamic_std = roll_alpha.rolling(window=window_size).std()
                up_range = dynamic_avg + dynamic_std
                dn_range = dynamic_avg - dynamic_std

                calc_table = pd.DataFrame({
                    f"{roll_n}Y Ann. Manager":    m_ann_roll.loc[roll_alpha.index],
                    f"{roll_n}Y Ann. Benchmark":  b_ann_roll.loc[roll_alpha.index],
                    f"{roll_n}Y Rolling Alpha":   roll_alpha,
                    "Dynamic Avg Alpha":           dynamic_avg,
                    "Dynamic Std Dev":             dynamic_std,
                    "Upper Range":                 up_range,
                    "Lower Range":                 dn_range,
                })

                st.write(f"**Complete Dynamic {roll_n}-Year Calculation Table**")
                st.dataframe(calc_table.iloc[::-1].style.format("{:.2%}"), use_container_width=True)

                fig_roll = go.Figure()
                fig_roll.add_trace(go.Scatter(
                    x=roll_alpha.index, y=roll_alpha,
                    name=f"{roll_n}Y Rolling Alpha",
                    line=dict(color="#002366", width=2.5)
                ))
                fig_roll.add_trace(go.Scatter(
                    x=up_range.index, y=up_range,
                    name="Upper Range",
                    line=dict(color="#800080", width=1.5, dash="dashdot")
                ))
                fig_roll.add_trace(go.Scatter(
                    x=dn_range.index, y=dn_range,
                    name="Lower Range",
                    line=dict(color="#008080", width=1.5, dash="dashdot")
                ))
                fig_roll.update_layout(
                    template="plotly_white",
                    title=f"{roll_n}-Year Rolling Alpha Profile: {alpha_mgr} vs {alpha_bench}",
                    yaxis_tickformat=".2%",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig_roll, use_container_width=True)
            else:
                st.warning(f"Requires at least {window_size} months of data for {roll_n}-Year calculation.")


# =====================================================
# 19. SIDEBAR EXPORT BUTTONS
# =====================================================
st.sidebar.divider()

# # CSV
# csv_data = pd.concat(results, axis=0).to_csv().encode("utf-8")
# st.sidebar.download_button("📄 Download CSV", data=csv_data,
#                             file_name=f"{dt.date.today()}_jarir_results.csv", mime="text/csv")

# =====================================================
# 19. SIDEBAR EXPORT BUTTONS
# =====================================================

st.sidebar.divider()


# ---- Format date properly ----
today_str = dt.datetime.today().strftime("%Y-%m-%d")


# =====================================================
# EXCEL EXPORT
# =====================================================
st.sidebar.divider()

if st.sidebar.button("📥 Generate Excel Report"):
    with st.spinner("Building Excel workbook..."):
        xl_buf = generate_excel(
            {m: results[m].apply(pd.to_numeric, errors='coerce') for m in metrics},
            manager_cols,
            cal_ret,
            cal_diff,
            alpha_disp_year,
            cap_tbl,
            wr_df
        )

        st.sidebar.download_button(
            "⬇️ Download Excel",
            data=xl_buf,
            file_name=f"{today_str}_{alpha_fund}_jarir_quant_analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# =====================================================
# PDF EXPORT
# =====================================================
st.sidebar.divider()

if st.sidebar.button("📑 Generate Master PDF"):
    with st.spinner("Building professional PDF report..."):

        pdf_bytes = generate_pdf_pretty(
            results_dict=results,
            mgrs=manager_cols,
            rfr_name=rfr_col,
            start_l=start_label,
            end_l=end_label,
            metrics_order=metrics,
            fig_rr_=fig_rr,
            fig_ud_=None,
            viz_h_=viz_h,
            alpha_bench=alpha_bench,
            alpha_fund=alpha_fund,
            #bench_diff=bench_diff
        )

        st.sidebar.download_button(
            "⬇️ Download PDF",
            data=pdf_bytes,
            file_name=f"{today_str}_{alpha_fund}_jarir_quant_analysis.pdf",
            mime="application/pdf"
        )